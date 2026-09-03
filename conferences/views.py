import json
from collections import Counter, defaultdict
from datetime import datetime

from django.db.models import Count
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render

from companymap.models import Company
from contactdb.models import Person

from .models import Conference, ConferenceCompany, Session, Speaker

DEFAULT_DURATION_MINUTES = 45
MIN_BLOCK_MINUTES = 20
AXIS_PADDING_MINUTES = 30
BANNER_THRESHOLD_MINUTES = 300


def home(request):
    conferences = Conference.objects.annotate(
        company_count=Count("companies", distinct=True),
        session_count=Count("sessions", distinct=True),
    )
    return render(request, "conferences/home.html", {"conferences": conferences})


def _minutes(t):
    return t.hour * 60 + t.minute


def _layout_day(sessions_with_times):
    """Greedy interval-column packing (same idea as Google Calendar's day
    view): overlapping sessions get placed side by side, non-overlapping
    ones share the full width. Returns [(session, col_index, col_count)]."""
    items = sorted(sessions_with_times, key=lambda x: x[1])
    clusters = []
    current, current_end = [], None
    for item in items:
        start, end = item[1], item[2]
        if current and start >= current_end:
            clusters.append(current)
            current, current_end = [item], end
        else:
            current.append(item)
            current_end = max(current_end or 0, end)
    if current:
        clusters.append(current)

    result = []
    for cluster in clusters:
        columns = []  # each entry: minute the column is free from
        placed = []
        for item in cluster:
            start, end = item[1], item[2]
            for ci, free_from in enumerate(columns):
                if start >= free_from:
                    columns[ci] = end
                    placed.append((item, ci))
                    break
            else:
                columns.append(end)
                placed.append((item, len(columns) - 1))
        total = len(columns)
        for item, ci in placed:
            result.append((item[0], ci, total))
    return result


def conference_detail(request, slug):
    conference = get_object_or_404(Conference, slug=slug)

    companies = conference.companies.select_related("company").all().order_by("name")
    tag_counts = Counter()
    for company in companies:
        for tag in company.tags:
            tag_counts[tag] += 1
    tag_labels = dict(ConferenceCompany.Tag.choices)
    tag_options = [
        {"value": value, "label": tag_labels.get(value, value), "count": count}
        for value, count in sorted(tag_counts.items(), key=lambda kv: -kv[1])
    ]

    UNMATCHED_SEGMENT = "__unmatched__"
    segment_counts = Counter()
    for company in companies:
        segment_counts[company.company.segment if (company.company and company.company.segment) else UNMATCHED_SEGMENT] += 1
    segment_options = [
        {"value": value, "label": "Not in our DB" if value == UNMATCHED_SEGMENT else value, "count": count}
        for value, count in sorted(segment_counts.items(), key=lambda kv: -kv[1])
    ]

    synthego_labels = dict(Company.SynthegoRelation.choices)
    for company in companies:
        company.data_tags = "|".join(company.tags)
        company.data_segment = company.company.segment if (company.company and company.company.segment) else UNMATCHED_SEGMENT
        company.tag_badges = [{"value": t, "label": tag_labels.get(t, t)} for t in company.tags]

        info_badges = []
        mapped = company.company
        if mapped:
            if mapped.segment:
                info_badges.append({"kind": "segment", "label": mapped.segment})
            if mapped.synthego_relation and mapped.synthego_relation != Company.SynthegoRelation.NO_RELATION:
                info_badges.append(
                    {"kind": "synthego_" + mapped.synthego_relation, "label": synthego_labels[mapped.synthego_relation]}
                )
            if mapped.has_marketing_presales_team:
                info_badges.append({"kind": "presales", "label": "Marketing/Presales"})
            if mapped.has_data_bi_team:
                info_badges.append({"kind": "data_bi", "label": "Data/BI"})
        company.info_badges = info_badges

    sessions = list(conference.sessions.all().order_by("day", "start_time", "id"))
    for session in sessions:
        session.data = json.dumps(
            {
                "title": session.title,
                "type": session.get_session_type_display(),
                "day": session.day.strftime("%A, %B %-d") if session.day else "",
                "time": _format_time_range(session),
                "location": session.location,
                "company": session.company_name,
                "speakers": session.speakers,
                "description": session.description,
            }
        )

    by_day = defaultdict(list)
    unscheduled_by_day = defaultdict(list)
    banner_by_day = defaultdict(list)
    all_minutes = []
    for session in sessions:
        if not session.day:
            continue
        if not session.start_time:
            unscheduled_by_day[session.day].append(session)
            continue
        start_min = _minutes(session.start_time)
        end_min = _minutes(session.end_time) if session.end_time else start_min + DEFAULT_DURATION_MINUTES
        if end_min <= start_min:
            end_min = start_min + MIN_BLOCK_MINUTES
        # Venue-wide "X open all day" logistics entries (Registration, Exhibit
        # Halls Open, Private Meeting Rooms...) span most of the day and would
        # otherwise force every real session into a narrow sliver via the
        # overlap-column packing below. Show those as a banner row instead,
        # like Google Calendar's all-day events.
        if session.session_type == Session.SessionType.LOGISTICS or (end_min - start_min) >= BANNER_THRESHOLD_MINUTES:
            banner_by_day[session.day].append(session)
            continue
        by_day[session.day].append((session, start_min, end_min))
        all_minutes.extend([start_min, end_min])

    if all_minutes:
        axis_start = max(0, min(all_minutes) - AXIS_PADDING_MINUTES)
        axis_end = min(24 * 60, max(all_minutes) + AXIS_PADDING_MINUTES)
    else:
        axis_start, axis_end = 8 * 60, 18 * 60
    axis_start = (axis_start // 60) * 60
    axis_end = ((axis_end + 59) // 60) * 60
    axis_span = max(axis_end - axis_start, 60)

    hour_marks = []
    hour = axis_start
    while hour <= axis_end:
        hour_marks.append(
            {
                "label": datetime(2000, 1, 1, (hour // 60) % 24, 0).strftime("%-I %p"),
                "top_pct": round((hour - axis_start) / axis_span * 100, 3),
            }
        )
        hour += 60

    days = []
    for day in sorted(by_day.keys() | unscheduled_by_day.keys() | banner_by_day.keys()):
        placed = _layout_day(by_day.get(day, []))
        blocks = []
        for session, ci, total in placed:
            start_min = _minutes(session.start_time)
            end_min = _minutes(session.end_time) if session.end_time else start_min + DEFAULT_DURATION_MINUTES
            if end_min <= start_min:
                end_min = start_min + MIN_BLOCK_MINUTES
            blocks.append(
                {
                    "session": session,
                    "top_pct": round((start_min - axis_start) / axis_span * 100, 3),
                    "height_pct": round((end_min - start_min) / axis_span * 100, 3),
                    "left_pct": round(ci / total * 100, 3),
                    "width_pct": round(100 / total, 3),
                }
            )
        days.append(
            {
                "date": day,
                "blocks": blocks,
                "unscheduled": unscheduled_by_day.get(day, []),
                "banners": banner_by_day.get(day, []),
            }
        )

    return render(
        request,
        "conferences/detail.html",
        {
            "conference": conference,
            "companies": companies,
            "tag_options": tag_options,
            "segment_options": segment_options,
            "days": days,
            "hour_marks": hour_marks,
            "total_sessions": len(sessions),
        },
    )


def _format_time_range(session):
    if session.start_time and session.end_time:
        return f"{session.start_time.strftime('%-I:%M %p')} – {session.end_time.strftime('%-I:%M %p')}"
    if session.start_time:
        return session.start_time.strftime("%-I:%M %p")
    return session.raw_time_label or "Time TBD"


def conference_export(request, slug):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    conference = get_object_or_404(Conference, slug=slug)
    attendees = list(conference.companies.select_related("company").all().order_by("name"))
    sessions = list(conference.sessions.all().order_by("day", "start_time", "id"))
    speakers = list(conference.speakers.all().order_by("name"))
    tag_labels = dict(ConferenceCompany.Tag.choices)

    linked_company_ids = [a.company_id for a in attendees if a.company_id]
    people = (
        Person.objects.filter(company_id__in=linked_company_ids)
        .select_related("company")
        .order_by("company__name", "name")
    )

    # Warm-connection roll-up per matched company: which connectors (if any)
    # have a prior relationship with someone there.
    warm_connectors_by_company = defaultdict(set)
    for p in people:
        if p.company_id and p.prior_connections:
            warm_connectors_by_company[p.company_id].update(p.prior_connections)

    # Session cross-reference: which session(s) is this attendee presenting,
    # matched on Session.company_name against the attendee's (and its mapped
    # Company's) name, case-insensitively.
    sessions_by_company_name = defaultdict(list)
    for s in sessions:
        key = (s.company_name or "").strip().lower()
        if key:
            sessions_by_company_name[key].append(s)

    def sessions_for_attendee(attendee):
        keys = {attendee.name.strip().lower()}
        if attendee.company_id:
            keys.add(attendee.company.name.strip().lower())
        found = []
        for key in keys:
            found.extend(sessions_by_company_name.get(key, []))
        # de-dupe while preserving order
        seen_ids = set()
        unique = []
        for s in found:
            if s.id not in seen_ids:
                seen_ids.add(s.id)
                unique.append(s)
        return unique

    wb = Workbook()
    header_font = Font(bold=True)

    def write_sheet(ws, headers, rows):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
        for row in rows:
            ws.append(row)
        widths = [len(str(h)) for h in headers]
        for row in rows:
            for i, val in enumerate(row):
                widths[i] = min(max(widths[i], len(str(val)) if val is not None else 0), 60)
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width + 2
        ws.freeze_panes = "A2"

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    write_sheet(
        ws,
        ["Field", "Value"],
        [
            ["Conference", conference.name],
            ["Location", conference.location],
            ["Start date", conference.start_date.isoformat() if conference.start_date else ""],
            ["End date", conference.end_date.isoformat() if conference.end_date else ""],
            ["Companies attending", len(attendees)],
            ["Companies matched in our DB", len(linked_company_ids)],
            ["People (from matched companies)", people.count()],
            ["Speakers", len(speakers)],
            ["Sessions", len(sessions)],
            ["Description", conference.description],
        ],
    )

    # --- Companies sheet ---
    ws = wb.create_sheet("Companies")
    rows = []
    for a in attendees:
        mapped = a.company
        warm = warm_connectors_by_company.get(a.company_id, set()) if a.company_id else set()
        attendee_sessions = sessions_for_attendee(a)
        session_summary = "; ".join(
            f"{s.title} — {s.day.strftime('%a') if s.day else ''} {_format_time_range(s)}".strip(" —")
            for s in attendee_sessions
        )
        rows.append([
            a.name,
            (mapped.domain if mapped and mapped.domain else "") or a.website,
            ", ".join(tag_labels.get(t, t) for t in a.tags),
            a.booth,
            a.website,
            a.linkedin_url,
            "Yes" if mapped else "No",
            mapped.segment if mapped else "",
            mapped.city if mapped else "",
            mapped.state_code if mapped else "",
            mapped.industry if mapped else "",
            mapped.account_owner if mapped else "",
            mapped.priority if mapped else "",
            "Yes" if (mapped and mapped.has_marketing_presales_team) else "",
            "Yes" if (mapped and mapped.has_data_bi_team) else "",
            "Yes" if warm else "",
            ", ".join(sorted(warm)),
            "Yes" if attendee_sessions else "",
            session_summary,
            a.notes,
        ])
    write_sheet(
        ws,
        ["Company Name", "Domain", "Tags", "Booth", "Website", "LinkedIn URL", "In Our DB",
         "Segment", "City", "State", "Industry", "Account Owner", "Priority",
         "Marketing/Presales Team", "Data/BI Team", "Warm Connection", "Warm Connectors",
         "Speaking?", "Session Details", "Notes"],
        rows,
    )

    # --- People sheet ---
    ws = wb.create_sheet("People")
    rows = []
    for p in people:
        rows.append([
            p.name,
            p.title,
            p.company.name if p.company_id else p.company_name,
            p.company.segment if p.company_id else "",
            p.email,
            p.linkedin_url,
            p.phone,
            p.segment,
            p.stakeholder_role,
            ", ".join(p.prior_connections),
        ])
    write_sheet(
        ws,
        ["Name", "Title", "Company", "Company Segment", "Email", "LinkedIn URL", "Phone",
         "Segment", "Stakeholder Role", "Prior Connections"],
        rows,
    )

    # --- Speakers sheet ---
    ws = wb.create_sheet("Speakers")
    rows = [
        [s.name, s.title, s.company_name, s.profile_url, s.photo_url]
        for s in speakers
    ]
    write_sheet(ws, ["Name", "Title", "Company", "Profile URL", "Photo URL"], rows)

    # --- Sessions sheet ---
    ws = wb.create_sheet("Sessions")
    rows = []
    for s in sessions:
        speaker_names = ", ".join(
            f"{sp.get('name', '')} ({sp.get('affiliation', '')})" if sp.get("affiliation") else sp.get("name", "")
            for sp in (s.speakers or [])
        )
        rows.append([
            s.day.isoformat() if s.day else "",
            s.start_time.strftime("%H:%M") if s.start_time else "",
            s.end_time.strftime("%H:%M") if s.end_time else "",
            s.raw_time_label,
            s.title,
            s.get_session_type_display(),
            s.location,
            s.company_name,
            speaker_names,
            s.description,
        ])
    write_sheet(
        ws,
        ["Day", "Start Time", "End Time", "Raw Time Label", "Title", "Type", "Location",
         "Presenting Company", "Speakers", "Description"],
        rows,
    )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"{conference.slug}-export.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
