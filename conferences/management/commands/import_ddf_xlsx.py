"""Import a conference workbook exported in the "DDF Summit" format: sheets
named 'Commercial Partners' (a flat sponsor/exhibitor list) and 'Agenda' (a
single tabular sheet with one row per session, already carrying structured
columns for time, speakers, and speaker companies). See conferences/models.py
for the resulting schema.
"""

import datetime
import re
from itertools import zip_longest
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from companymap.models import Company
from conferences.management.commands.import_conference_xlsx import (
    canon,
    clean_text,
    parse_time_range,
    _merge_duplicate_companies,
)
from conferences.models import Conference, ConferenceCompany, Session, Speaker

SESSION_TYPE_MAP = {
    "registration": Session.SessionType.LOGISTICS,
    "opening": Session.SessionType.OTHER,
    "keynote": Session.SessionType.KEYNOTE,
    "case study": Session.SessionType.OTHER,
    "solution spotlight": Session.SessionType.EXHIBITOR_PRESENTATION,
    "refreshment break": Session.SessionType.LOGISTICS,
    "lunch": Session.SessionType.MEAL,
    "close of day": Session.SessionType.OTHER,
}

SOLUTION_SPOTLIGHT_BY_RE = re.compile(r"^Solution Spotlight by (.+)$", re.IGNORECASE)


def clean_company_list(text):
    """'| Biogen | | Johnson and Johnson' -> ['Biogen', 'Johnson and Johnson']"""
    if not text:
        return []
    parts = [clean_text(re.sub(r"^,\s*", "", p.strip())) for p in str(text).split("|")]
    seen = set()
    result = []
    for part in parts:
        if part and part not in seen:
            seen.add(part)
            result.append(part)
    return result


class Command(BaseCommand):
    help = "Import a DDF-Summit-format conference workbook (.xlsx) into the conferences app."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str)
        parser.add_argument("--name", required=True)
        parser.add_argument("--location", default="")
        parser.add_argument("--start-date", default="", help="YYYY-MM-DD")
        parser.add_argument("--end-date", default="", help="YYYY-MM-DD")

    @transaction.atomic
    def handle(self, *args, **options):
        path = Path(options["xlsx_path"])
        if not path.exists():
            raise CommandError(f"File not found: {path}")

        wb = openpyxl.load_workbook(path, data_only=True)

        start_date = (
            datetime.date.fromisoformat(options["start_date"]) if options["start_date"] else None
        )

        conference, created = Conference.objects.update_or_create(
            name=options["name"],
            defaults={
                "location": options["location"],
                "start_date": start_date,
                "end_date": options["end_date"] or None,
                "source_file": path.name,
            },
        )
        self.stdout.write(f"{'Created' if created else 'Updated'} conference {conference.name!r}")

        conference.companies.all().delete()
        conference.sessions.all().delete()
        conference.speakers.all().delete()

        companies = {}

        def stage_company(raw_name, tag=None, notes=""):
            name = clean_text(raw_name)
            if not name:
                return None
            key = canon(name)
            if not key:
                return None
            entry = companies.get(key)
            if entry is None:
                entry = {"name": name, "tags": set(), "booth": "", "website": "",
                          "linkedin_url": "", "notes": []}
                companies[key] = entry
            if tag:
                entry["tags"].add(tag)
            if notes:
                entry["notes"].append(notes)
            return entry

        partner_count = self._import_commercial_partners(wb, stage_company)
        session_count, speaker_count = self._import_agenda(wb, conference, stage_company, start_date)

        _merge_duplicate_companies(companies)

        company_ids_by_name = {c.name.lower(): c.id for c in Company.objects.only("id", "name")}
        created_companies = []
        for entry in companies.values():
            matched = company_ids_by_name.get(entry["name"].lower())
            created_companies.append(
                ConferenceCompany(
                    conference=conference,
                    name=entry["name"],
                    tags=sorted(entry["tags"]),
                    booth=entry["booth"],
                    website=entry["website"],
                    linkedin_url=entry["linkedin_url"],
                    notes="\n".join(entry["notes"]),
                    company_id=matched,
                )
            )
        ConferenceCompany.objects.bulk_create(created_companies)

        self.stdout.write(self.style.SUCCESS(
            f"Companies: {len(created_companies)} (partners {partner_count}) | "
            f"Speakers: {speaker_count} | Sessions: {session_count}"
        ))

    def _sheet(self, wb, *candidates):
        for name in candidates:
            if name in wb.sheetnames:
                return wb[name]
        raise CommandError(f"None of {candidates} found. Sheets: {wb.sheetnames}")

    def _import_commercial_partners(self, wb, stage_company):
        ws = self._sheet(wb, "Commercial Partners")
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[1] if len(row) > 1 else None
            if not name:
                continue
            profile_url = row[2] if len(row) > 2 else ""
            notes = f"DDF Summit profile: {profile_url}" if profile_url else ""
            stage_company(name, tag=ConferenceCompany.Tag.EXHIBITOR, notes=notes)
            count += 1
        return count

    def _import_agenda(self, wb, conference, stage_company, start_date):
        ws = self._sheet(wb, "Agenda")
        sessions = []
        speakers_seen = {}
        day_offset = 0
        prev_time = None

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[4] is None and row[6] is None:
                continue
            title = clean_text(row[4])
            raw_type = clean_text(row[5]).lower()
            time_value = row[6]
            room = clean_text(row[8]) if len(row) > 8 else ""
            speaker_count = row[10] or 0
            speaker_names = clean_text(row[11])
            speaker_roles = clean_text(row[12])
            speaker_companies_raw = row[13]
            description = clean_text(row[16]) if len(row) > 16 else ""

            if isinstance(time_value, datetime.time):
                if prev_time is not None and time_value < prev_time:
                    day_offset += 1
                prev_time = time_value

            day = None
            if start_date:
                day = start_date + datetime.timedelta(days=day_offset)

            time_range_text = row[7] if len(row) > 7 else None
            start_time, end_time, raw_label = parse_time_range(time_range_text)
            if not start_time and isinstance(time_value, datetime.time):
                start_time = time_value

            if not raw_type:
                session_type = (
                    Session.SessionType.NETWORKING
                    if title and "networking" in title.lower()
                    else Session.SessionType.OTHER
                )
            else:
                session_type = SESSION_TYPE_MAP.get(raw_type, Session.SessionType.OTHER)

            companies_for_row = clean_company_list(speaker_companies_raw)

            company_name = ""
            if session_type == Session.SessionType.EXHIBITOR_PRESENTATION:
                if companies_for_row:
                    company_name = companies_for_row[0]
                else:
                    match = SOLUTION_SPOTLIGHT_BY_RE.match(title or "")
                    if match:
                        company_name = clean_text(match.group(1))
                if company_name:
                    stage_company(company_name, tag=ConferenceCompany.Tag.EXHIBITOR)
            elif companies_for_row:
                for company in companies_for_row:
                    stage_company(company, tag=ConferenceCompany.Tag.SPEAKER_SESSION)

            names_list = [clean_text(n) for n in speaker_names.split("|")] if speaker_names else []
            speakers_json = [
                {"name": name, "affiliation": company}
                for name, company in zip_longest(names_list, companies_for_row, fillvalue="")
                if name or company
            ]
            for entry in speakers_json:
                if entry["name"] and entry["name"] not in speakers_seen:
                    speakers_seen[entry["name"]] = entry["affiliation"]

            sessions.append(
                Session(
                    conference=conference,
                    title=title or (session_type or "Session"),
                    session_type=session_type,
                    day=day,
                    start_time=start_time,
                    end_time=end_time,
                    raw_time_label=raw_label or clean_text(time_range_text),
                    description=description,
                    speakers=speakers_json,
                    company_name=company_name,
                    location=room,
                    source_sheet="Agenda",
                )
            )

        Session.objects.bulk_create(sessions)

        speaker_records = [
            Speaker(conference=conference, name=name, company_name=affiliation)
            for name, affiliation in speakers_seen.items()
        ]
        Speaker.objects.bulk_create(speaker_records)

        return len(sessions), len(speaker_records)
