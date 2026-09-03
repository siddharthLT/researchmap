"""Import a conference workbook exported in the "Contract Pharma" format:
sheets named '<Slug>_Speakers' (one row per talk, with a combined
"Name, Title, Company" cell), '<Slug>_Sponsors' (flat sponsor list with a
tier), '<Slug>_Exhibitors' (flat list with booth numbers), and an optional
'Linkedin Capture' sheet (ignored — no LinkedIn capture yet for this
format). See conferences/models.py for the resulting schema.
"""

import datetime
import re
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from companymap.models import Company
from conferences.management.commands.import_conference_xlsx import (
    canon,
    classify_session_type,
    clean_text as _clean_text_base,
    parse_time_range,
    _merge_duplicate_companies,
)
from conferences.models import Conference, ConferenceCompany, Session, Speaker

MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dec": 12,
}
DATE_RE = re.compile(r"([a-zA-Z]+)\.?\s+(\d{1,2})")


def clean_text(value):
    """Same as the ChemOutsourcing clean_text, plus non-breaking-space normalization."""
    text = _clean_text_base(value)
    return text.replace("\xa0", " ").strip()


def parse_hour_cell(value):
    """Speakers sheet time/end-time cells are usually a datetime.time, but a
    handful are a bare number (e.g. 9.0, 10.0) meaning that hour on the dot."""
    if isinstance(value, datetime.time):
        return value
    if isinstance(value, (int, float)):
        hour = int(value)
        if 0 <= hour <= 23:
            return datetime.time(hour, 0)
    return None


def parse_event_date(text, year):
    if not text or not year:
        return None
    match = DATE_RE.search(str(text).strip().lower())
    if not match:
        return None
    month = MONTHS.get(match.group(1).strip().rstrip("."))
    if not month:
        return None
    try:
        return datetime.date(year, month, int(match.group(2)))
    except ValueError:
        return None


def parse_role_company(speaker_name, role_company):
    """'Name, Title bits..., Company' -> (title, company). The leading name is
    sometimes omitted, and sometimes doesn't exactly match speaker_name (extra
    periods, etc.), so strip it only on a loose (period/space-insensitive)
    match rather than requiring an exact one."""
    text = clean_text(role_company)
    if not text:
        return "", ""
    parts = [p.strip() for p in text.split(",") if p.strip()]
    if not parts:
        return "", ""

    def loose(s):
        return re.sub(r"[.\s]+", " ", s).strip().lower()

    if loose(parts[0]) == loose(speaker_name):
        parts = parts[1:]
    if not parts:
        return "", ""
    if len(parts) == 1:
        return "", parts[0]
    return ", ".join(parts[:-1]), parts[-1]


class Command(BaseCommand):
    help = "Import a Contract-Pharma-format conference workbook (.xlsx) into the conferences app."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str)
        parser.add_argument("--name", required=True)
        parser.add_argument("--location", default="")
        parser.add_argument("--start-date", default="", help="YYYY-MM-DD")
        parser.add_argument("--end-date", default="", help="YYYY-MM-DD")
        parser.add_argument("--sheet-prefix", default="", help="e.g. 'ContractPharma_2026' if sheet names don't match --name")

    @transaction.atomic
    def handle(self, *args, **options):
        path = Path(options["xlsx_path"])
        if not path.exists():
            raise CommandError(f"File not found: {path}")

        wb = openpyxl.load_workbook(path, data_only=True)
        prefix = options["sheet_prefix"] or self._guess_prefix(wb)
        year = int(options["start_date"][:4]) if options["start_date"] else None

        conference, created = Conference.objects.update_or_create(
            name=options["name"],
            defaults={
                "location": options["location"],
                "start_date": options["start_date"] or None,
                "end_date": options["end_date"] or None,
                "source_file": path.name,
            },
        )
        self.stdout.write(f"{'Created' if created else 'Updated'} conference {conference.name!r}")

        conference.companies.all().delete()
        conference.sessions.all().delete()
        conference.speakers.all().delete()

        companies = {}

        def stage_company(raw_name, tag=None, booth="", notes=""):
            name = clean_text(raw_name)
            if not name:
                return None
            key = canon(name)
            if not key:
                return None
            entry = companies.get(key)
            if entry is None:
                entry = {"name": name, "tags": set(), "booth": booth, "website": "",
                          "linkedin_url": "", "notes": []}
                companies[key] = entry
            if tag:
                entry["tags"].add(tag)
            if booth and not entry["booth"]:
                entry["booth"] = booth
            if notes:
                entry["notes"].append(notes)
            return entry

        exhibitor_count = self._import_exhibitors(wb, prefix, stage_company)
        sponsor_count = self._import_sponsors(wb, prefix, stage_company)
        speaker_count, session_count = self._import_speakers(wb, prefix, conference, stage_company, year)

        _merge_duplicate_companies(companies)

        # --- company matching against companymap.Company, using the same
        # loose canon() key as staging (far more forgiving than an exact
        # lowercase name match: "Alcami" / "Alcami Corporation" both resolve
        # to the same key) ---
        existing_by_canon = {}
        for c in Company.objects.only("id", "name"):
            key = canon(c.name)
            if key and key not in existing_by_canon:
                existing_by_canon[key] = c.id

        created_companies = []
        unmatched = []
        matched_count = 0
        for key, entry in companies.items():
            matched_id = existing_by_canon.get(key)
            if matched_id:
                matched_count += 1
            else:
                unmatched.append(entry["name"])
            created_companies.append(
                ConferenceCompany(
                    conference=conference,
                    name=entry["name"],
                    tags=sorted(entry["tags"]),
                    booth=entry["booth"],
                    website=entry["website"],
                    linkedin_url=entry["linkedin_url"],
                    notes="\n".join(entry["notes"]),
                    company_id=matched_id,
                )
            )
        ConferenceCompany.objects.bulk_create(created_companies)

        self.stdout.write(self.style.SUCCESS(
            f"Companies: {len(created_companies)} (exhibitors {exhibitor_count}, sponsors {sponsor_count}, "
            f"matched to DB {matched_count}, unmatched {len(unmatched)}) | "
            f"Speakers: {speaker_count} | Sessions: {session_count}"
        ))
        if unmatched:
            self.stdout.write("Unmatched company names (not yet in companymap.Company):")
            for name in sorted(unmatched):
                self.stdout.write(f"  - {name}")

    def _guess_prefix(self, wb):
        for name in wb.sheetnames:
            if name.endswith("_Exhibitors"):
                return name[: -len("_Exhibitors")]
        return ""

    def _sheet(self, wb, name):
        if name not in wb.sheetnames:
            raise CommandError(f"Sheet {name!r} not found. Have: {wb.sheetnames}")
        return wb[name]

    def _import_exhibitors(self, wb, prefix, stage_company):
        ws = self._sheet(wb, f"{prefix}_Exhibitors")
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[1] if len(row) > 1 else None
            if not name:
                continue
            booth = row[2] if len(row) > 2 else None
            booth_str = str(int(booth)) if isinstance(booth, (int, float)) else clean_text(booth)
            stage_company(name, tag=ConferenceCompany.Tag.EXHIBITOR, booth=booth_str)
            count += 1
        return count

    def _import_sponsors(self, wb, prefix, stage_company):
        ws = self._sheet(wb, f"{prefix}_Sponsors")
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[1] if len(row) > 1 else None
            if not name:
                continue
            tier = clean_text(row[2]) if len(row) > 2 else ""
            stage_company(
                name,
                tag=ConferenceCompany.Tag.EXHIBITOR,
                notes=f"Sponsor tier: {tier}" if tier else "Sponsor",
            )
            count += 1
        return count

    def _import_speakers(self, wb, prefix, conference, stage_company, year):
        ws = self._sheet(wb, f"{prefix}_Speakers")
        speakers = []
        sessions = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            speaker_name = clean_text(row[1]) if len(row) > 1 else ""
            if not speaker_name:
                continue
            role_company = row[2] if len(row) > 2 else ""
            session_title = clean_text(row[3]) if len(row) > 3 else ""
            profile_url = clean_text(row[4]) if len(row) > 4 else ""
            start_time = parse_hour_cell(row[5]) if len(row) > 5 else None
            end_time = parse_hour_cell(row[6]) if len(row) > 6 else None
            date_text = row[7] if len(row) > 7 else ""
            day = parse_event_date(date_text, year)

            title, company_name = parse_role_company(speaker_name, role_company)

            speakers.append(Speaker(
                conference=conference,
                name=speaker_name,
                title=title,
                company_name=company_name,
                profile_url=profile_url,
            ))
            if company_name:
                stage_company(company_name, tag=ConferenceCompany.Tag.SPEAKER_SESSION)

            sessions.append(Session(
                conference=conference,
                title=session_title or f"{speaker_name} talk",
                session_type=classify_session_type(session_title or ""),
                day=day,
                start_time=start_time,
                end_time=end_time,
                raw_time_label=clean_text(date_text),
                speakers=[{"name": speaker_name, "affiliation": company_name}],
                source_sheet="Speakers",
            ))

        Speaker.objects.bulk_create(speakers)
        Session.objects.bulk_create(sessions)
        return len(speakers), len(sessions)
