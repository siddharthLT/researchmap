"""Import a conference workbook exported in the "ChemOutsourcing" format:
sheets named 'Attendees list', 'Agenda', 'Speakers List', 'Exhibitors
Presentation', 'Participating Speed Networking', and (optional) 'Linkedin
Post capture'. See conferences/models.py for the resulting schema.
"""

import datetime
import difflib
import re
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from companymap.models import Company
from conferences.models import Conference, ConferenceCompany, Session, Speaker

MOJIBAKE_FIXES = {
    "â€“": "–",
    "â€”": "—",
    "â€™": "’",
    "â€˜": "‘",
    "â€œ": "“",
    "â€\x9d": "”",
    "Ã©": "é",
}

WEEKDAY_LABEL_RE = re.compile(r"^(Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday),?$", re.IGNORECASE)
CONT_LABEL_RE = re.compile(r"\(cont\)", re.IGNORECASE)
TIME_RANGE_RE = re.compile(
    r"^\s*(\d{1,2}:\d{2}\s*(?:AM|PM)?)\s*[–—-]\s*(\d{1,2}:\d{2}\s*(?:AM|PM)?)\s*$", re.IGNORECASE
)
TRAILING_PAREN_RE = re.compile(r"^(.*?)\s*\(([^)]*)\)\s*$")

STOPWORDS = re.compile(
    r"\b(inc|incorporated|llc|l l c|ltd|limited|corp|corporation|co|company|"
    r"sa|s a|sl|s l|srl|s r l|gmbh|group|pharmaceuticals|pharmaceutical|pharma)\b"
)


def clean_text(value):
    if not value:
        return ""
    text = str(value).strip()
    for bad, good in MOJIBAKE_FIXES.items():
        text = text.replace(bad, good)
    return text


def canon(name):
    s = (name or "").lower()
    s = re.sub(r"\([^)]*\)", " ", s)
    s = re.sub(r"[.,&'\"/\-]", " ", s)
    s = STOPWORDS.sub(" ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def parse_time_range(label):
    """Return (start_time, end_time, raw_label) from a string or datetime.time cell."""
    if isinstance(label, datetime.time):
        return label, None, ""
    if isinstance(label, datetime.datetime):
        return label.time(), None, ""
    text = clean_text(label)
    if not text:
        return None, None, ""
    match = TIME_RANGE_RE.match(text)
    if not match:
        return None, None, text
    start_raw, end_raw = match.groups()
    start = _parse_clock(start_raw, end_raw)
    end = _parse_clock(end_raw, end_raw)
    return start, end, text if not (start and end) else ""


def _parse_clock(raw, fallback_meridiem_source):
    raw = raw.strip().upper()
    if "AM" not in raw and "PM" not in raw:
        # Borrow AM/PM from the other side of the range (e.g. "10:15 – 10:30 AM"
        # implies 10:15 AM too). Otherwise default to AM, except a bare "12:xx"
        # which almost always means noon in this context, not midnight -- the
        # caller (_import_exhibitor_presentations) also corrects any later
        # entries that would otherwise run backwards across the noon boundary.
        if "PM" in fallback_meridiem_source.upper():
            meridiem = "PM"
        elif raw.split(":")[0].strip() == "12":
            meridiem = "PM"
        else:
            meridiem = "AM"
        raw = f"{raw} {meridiem}"
    for fmt in ("%I:%M %p", "%I:%M%p"):
        try:
            return datetime.datetime.strptime(raw, fmt).time()
        except ValueError:
            continue
    return None


def classify_session_type(title):
    t = title.lower()
    if "workshop" in t:
        return Session.SessionType.WORKSHOP
    if "panel discussion" in t:
        return Session.SessionType.PANEL
    if "keynote" in t:
        return Session.SessionType.KEYNOTE
    if "speed networking" in t or "networking" in t:
        return Session.SessionType.NETWORKING
    if any(w in t for w in ("luncheon", "reception", "party", "drinks", "dinner", "hors d")):
        return Session.SessionType.MEAL
    if any(
        w in t
        for w in (
            "registration", "badge", "exhibit halls open", "private meeting rooms",
            "exhibitor set up", "exhibitor teardown", "event conclusion",
        )
    ):
        return Session.SessionType.LOGISTICS
    return Session.SessionType.OTHER


def parse_speaker_credits(text):
    """Split embedded 'Name – Company' lines (following a blank line after the
    title) into structured speaker dicts."""
    parts = text.split("\n\n", 1)
    title = parts[0].strip()
    speakers = []
    if len(parts) > 1:
        for line in parts[1].splitlines():
            line = line.strip()
            if not line:
                continue
            bits = re.split(r"\s+[–—-]\s+", line, maxsplit=1)
            if len(bits) == 2:
                speakers.append({"name": bits[0].strip(), "affiliation": bits[1].strip()})
            else:
                speakers.append({"name": line, "affiliation": ""})
    return title, speakers


def _merge_into(companies, primary_key, other_key):
    if primary_key == other_key or other_key not in companies or primary_key not in companies:
        return
    primary = companies[primary_key]
    other = companies.pop(other_key)
    primary["tags"] |= other["tags"]
    if not primary["booth"]:
        primary["booth"] = other["booth"]
    if not primary["website"]:
        primary["website"] = other["website"]
    if not primary["linkedin_url"]:
        primary["linkedin_url"] = other["linkedin_url"]
    primary["notes"].extend(other["notes"])


def _merge_duplicate_companies(companies):
    """Two independent passes to fold together entries the canonical-name key
    alone doesn't catch: same booth number (strongest signal), or a near-
    identical / prefix-matching canonical name (e.g. "CordenPharma" vs
    "Corden Pharma", "Moehs" vs "Moehs Iberica")."""
    by_booth = {}
    for key, entry in companies.items():
        if entry["booth"]:
            by_booth.setdefault(entry["booth"], []).append(key)
    for keys in by_booth.values():
        if len(keys) < 2:
            continue
        keys = sorted((k for k in keys if k in companies), key=lambda k: -len(companies[k]["name"]))
        for other_key in keys[1:]:
            _merge_into(companies, keys[0], other_key)

    keys = list(companies.keys())
    for i, a in enumerate(keys):
        if a not in companies:
            continue
        for b in keys[i + 1:]:
            if b not in companies or a not in companies:
                continue
            ratio = difflib.SequenceMatcher(None, a, b).ratio()
            is_prefix_match = (a in b or b in a) and min(len(a), len(b)) >= 4
            if ratio >= 0.88 or is_prefix_match:
                if len(companies[a]["name"]) >= len(companies[b]["name"]):
                    _merge_into(companies, a, b)
                else:
                    _merge_into(companies, b, a)


class Command(BaseCommand):
    help = "Import a ChemOutsourcing-format conference workbook (.xlsx) into the conferences app."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str)
        parser.add_argument("--name", required=True)
        parser.add_argument("--location", default="")
        parser.add_argument("--start-date", default="", help="YYYY-MM-DD")
        parser.add_argument("--end-date", default="", help="YYYY-MM-DD")

    @transaction.atomic
    def handle(self, *args, **options):
        path = Path(options["xlsx_path"])
        if not path.exists():
            raise CommandError(f"File not found: {path}")

        wb = openpyxl.load_workbook(path, data_only=True)

        conference, created = Conference.objects.update_or_create(
            name=options["name"],
            defaults={
                "location": options["location"],
                "start_date": options["start_date"] or None,
                "end_date": options["end_date"] or None,
                "source_file": path.name,
            },
        )
        self.stdout.write(f"{'Created' if created else 'Updated'} conference {conference.name!r}")

        # Wipe and rebuild this conference's derived data so re-imports are clean.
        conference.companies.all().delete()
        conference.sessions.all().delete()
        conference.speakers.all().delete()

        companies = {}  # canon name -> ConferenceCompany (unsaved, staged)

        def stage_company(raw_name, tag=None, booth="", website="", linkedin_url="", notes=""):
            name = clean_text(raw_name)
            if not name:
                return None
            key = canon(name)
            if not key:
                return None
            entry = companies.get(key)
            if entry is None:
                entry = {"name": name, "tags": set(), "booth": booth, "website": website,
                          "linkedin_url": linkedin_url, "notes": []}
                companies[key] = entry
            if tag:
                entry["tags"].add(tag)
            if booth and not entry["booth"]:
                entry["booth"] = booth
            if website and not entry["website"]:
                entry["website"] = website
            if linkedin_url and not entry["linkedin_url"]:
                entry["linkedin_url"] = linkedin_url
            if notes:
                entry["notes"].append(notes)
            return entry

        self._import_attendees(wb, stage_company)
        self._import_speed_networking(wb, stage_company)
        speaker_count = self._import_speakers(wb, conference, stage_company)
        session_count_agenda = self._import_agenda(wb, conference, stage_company)
        session_count_exhibitor = self._import_exhibitor_presentations(wb, conference, stage_company)
        self._import_linkedin_posts(wb, stage_company, companies)

        _merge_duplicate_companies(companies)
        for entry in companies.values():
            if entry["booth"]:
                entry["tags"].add(ConferenceCompany.Tag.EXHIBITOR)

        company_ids_by_name = {c.name.lower(): c.id for c in Company.objects.only("id", "name")}
        created_companies = []
        for entry in companies.values():
            matched = company_ids_by_name.get(entry["name"].lower())
            created_companies.append(
                ConferenceCompany(
                    conference=conference,
                    name=entry["name"],
                    tags=sorted(entry["tags"]),
                    booth=entry["booth"],
                    website=entry["website"],
                    linkedin_url=entry["linkedin_url"],
                    notes="\n".join(entry["notes"]),
                    company_id=matched,
                )
            )
        ConferenceCompany.objects.bulk_create(created_companies)

        self.stdout.write(self.style.SUCCESS(
            f"Companies: {len(created_companies)} | Speakers: {speaker_count} | "
            f"Sessions: {session_count_agenda + session_count_exhibitor} "
            f"(agenda {session_count_agenda}, exhibitor {session_count_exhibitor})"
        ))

    def _sheet(self, wb, *candidates):
        for name in candidates:
            if name in wb.sheetnames:
                return wb[name]
        for actual in wb.sheetnames:
            if actual.strip() in [c.strip() for c in candidates]:
                return wb[actual]
        raise CommandError(f"None of these sheets found: {candidates}. Have: {wb.sheetnames}")

    def _import_attendees(self, wb, stage_company):
        ws = self._sheet(wb, "Attendees list")
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw = clean_text(row[0])
            if not raw:
                continue
            match = TRAILING_PAREN_RE.match(raw)
            name, booth = (match.group(1), match.group(2)) if match else (raw, "")
            stage_company(name, booth=booth)

    def _import_speed_networking(self, wb, stage_company):
        ws = self._sheet(wb, "Participating Speed Networking")
        for row in ws.iter_rows(min_row=3, values_only=True):
            name = clean_text(row[0]) if len(row) > 0 else ""
            if not name:
                continue
            linkedin_url = clean_text(row[1]) if len(row) > 1 else ""
            website = clean_text(row[2]) if len(row) > 2 else ""
            stage_company(
                name,
                tag=ConferenceCompany.Tag.NETWORKING_SESSION,
                website=website,
                linkedin_url=linkedin_url,
            )

    def _import_speakers(self, wb, conference, stage_company):
        ws = self._sheet(wb, "Speakers List")
        speakers = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            name = clean_text(row[0]) if len(row) > 0 else ""
            if not name:
                continue
            title = clean_text(row[1]) if len(row) > 1 else ""
            company_name = clean_text(row[2]) if len(row) > 2 else ""
            profile_url = clean_text(row[3]) if len(row) > 3 else ""
            photo_url = clean_text(row[4]) if len(row) > 4 else ""
            speakers.append(
                Speaker(
                    conference=conference,
                    name=name,
                    title=title,
                    company_name=company_name,
                    profile_url=profile_url,
                    photo_url=photo_url,
                )
            )
            if company_name:
                stage_company(company_name, tag=ConferenceCompany.Tag.SPEAKER_SESSION)
        Speaker.objects.bulk_create(speakers)
        return len(speakers)

    def _import_agenda(self, wb, conference, stage_company):
        ws = self._sheet(wb, "Agenda ", "Agenda")
        sessions = []
        current_day = None
        for row in ws.iter_rows(min_row=1, values_only=True):
            cell_a, cell_b = row[0], row[1] if len(row) > 1 else None
            if cell_a is None:
                continue
            if isinstance(cell_a, datetime.datetime):
                current_day = cell_a.date()
                continue
            text = clean_text(cell_a)
            if not text or text in ("Agenda",) or WEEKDAY_LABEL_RE.match(text):
                continue
            if text.lower().startswith("explore keynote"):
                continue
            if text.lower().startswith("exhibitor presentations"):
                continue  # real sessions come from the Exhibitors Presentation sheet

            title, speaker_credits = parse_speaker_credits(text)
            start_time, end_time, raw_label = parse_time_range(cell_b)
            sessions.append(
                Session(
                    conference=conference,
                    title=title,
                    session_type=classify_session_type(title),
                    day=current_day,
                    start_time=start_time,
                    end_time=end_time,
                    raw_time_label=raw_label,
                    speakers=speaker_credits,
                    source_sheet="Agenda",
                )
            )
            for credit in speaker_credits:
                if credit["affiliation"]:
                    stage_company(credit["affiliation"], tag=ConferenceCompany.Tag.SPEAKER_SESSION)
        Session.objects.bulk_create(sessions)
        return len(sessions)

    def _import_exhibitor_presentations(self, wb, conference, stage_company):
        ws = self._sheet(wb, "Exhibitors Presentation ", "Exhibitors Presentation")
        sessions = []
        current_day = None
        buffer = []
        last_end_minutes = None

        def to_minutes(t):
            return t.hour * 60 + t.minute

        def flush(time_value):
            nonlocal last_end_minutes
            if not buffer:
                return
            company_name = buffer[-1]
            title = buffer[-2] if len(buffer) >= 2 else ""
            start_time, end_time, raw_label = parse_time_range(time_value)

            # These slot times never carry AM/PM. Since presentations run in
            # sequence through the day, if a slot would otherwise appear to
            # start before the previous one ended, it's actually the PM half
            # of the 12-hour clock.
            if start_time and last_end_minutes is not None and to_minutes(start_time) < last_end_minutes - 5:
                start_time = datetime.time((start_time.hour + 12) % 24, start_time.minute)
                if end_time:
                    end_time = datetime.time((end_time.hour + 12) % 24, end_time.minute)
            if end_time:
                last_end_minutes = to_minutes(end_time)
            elif start_time:
                last_end_minutes = to_minutes(start_time)

            sessions.append(
                Session(
                    conference=conference,
                    title=title or f"{company_name} presentation",
                    session_type=Session.SessionType.EXHIBITOR_PRESENTATION,
                    day=current_day,
                    start_time=start_time,
                    end_time=end_time,
                    raw_time_label=raw_label,
                    company_name=company_name,
                    source_sheet="Exhibitors Presentation",
                )
            )
            if company_name.strip().lower() != "open slot":
                stage_company(company_name, tag=ConferenceCompany.Tag.EXHIBITOR)
            buffer.clear()

        for row in ws.iter_rows(min_row=1, values_only=True):
            value = row[0]
            if value is None:
                continue
            if isinstance(value, datetime.datetime):
                flush(None)
                current_day = value.date()
                last_end_minutes = None
                continue
            text = clean_text(value)
            if not text or text == "Exhibitor's Presentation":
                continue
            if WEEKDAY_LABEL_RE.match(text) or CONT_LABEL_RE.search(text):
                # Page-break marker: flush whatever's buffered (a title+company
                # pair with no trailing time, e.g. the last slot before a page
                # break) rather than letting it bleed into the next block.
                flush(None)
                continue
            if TIME_RANGE_RE.match(text):
                flush(text)
                continue
            buffer.append(text)
        flush(None)

        Session.objects.bulk_create(sessions)
        return len(sessions)

    def _import_linkedin_posts(self, wb, stage_company, companies):
        try:
            ws = self._sheet(wb, "Linkedin Post capture ", "Linkedin Post capture")
        except CommandError:
            return
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            company_raw = clean_text(row[2]) if len(row) > 2 else ""
            if not company_raw:
                continue
            booth = row[10] if len(row) > 10 else None
            booth_str = str(int(booth)) if isinstance(booth, (int, float)) else ""
            linkedin_page = clean_text(row[7]) if len(row) > 7 else ""
            attendee_names = {
                clean_text(row[i]) for i in (13, 15, 17) if len(row) > i and row[i]
            }
            attendee1 = clean_text(row[13]) if len(row) > 13 else ""
            attendee1_title = clean_text(row[14]) if len(row) > 14 else ""
            note = f"{attendee1} ({attendee1_title})" if attendee1 else ""

            # "Company (Original Poster)" is sometimes just the individual's own
            # name (a personal post, not a company page) -- never stage that as
            # a company; if it has a booth, fold the note into whichever real
            # company already claims that booth instead.
            if company_raw in attendee_names:
                if booth_str:
                    target = next(
                        (e for e in companies.values() if e["booth"] == booth_str), None
                    )
                    if target:
                        target["tags"].add(ConferenceCompany.Tag.EXHIBITOR)
                        if note:
                            target["notes"].append(note)
                continue

            key = canon(company_raw)
            if key not in companies and not booth_str:
                continue  # no booth signal and not already a known company -- likely a media name
            stage_company(
                company_raw,
                tag=ConferenceCompany.Tag.EXHIBITOR if booth_str else None,
                booth=booth_str,
                linkedin_url=linkedin_page,
                notes=note,
            )
